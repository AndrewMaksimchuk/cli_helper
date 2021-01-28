# 
# 
import os
from openpyxl import *
from excel_styles import *
from create_excel_file_templates import *
import tools


def create_workbook_list(path_to_this_file, helps, data = None, 
										file_name = "excel_book_simple_labels", 
										column_number = 4,
										labels_width = 21,
										row_height = 15,
										font_size_name = 11, 
										font_size_code = 11, 
										font_size_price = 11, 
										alignment_name_horizontal = "left", 
										alignment_name_vertical = "center", 
										alignment_code_horizontal = "left", 
										alignment_code_vertical = "center", 
										alignment_price_horizontal = "left", 
										alignment_price_vertical = "center", 
										):
		if (not data):
			return None
		wb = Workbook()
		ws = wb.create_sheet("book", 0)
		book = wb["book"]

		CODE = "Код: "
		PRICE = "Ціна, грн: "

		# Зміні які зберігають позицію по рядку і колонці для переміщення
		row_mini = 2
		column = 1

		def header(header_text = "", text_size = 18, alignment_text_horizontal ="center", alignment_text_vertical = "center", header_height = False):
			book.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(column_order))
			book.cell(1, column = 1, value = str(header_text))
			book.cell(1, column = 1).font = styles.Font(color=styles.colors.WHITE, bold=True, size = text_size)
			book.cell(1, column = 1).fill = background_color_name
			book.cell(1, column = 1).alignment = styles.Alignment(horizontal=alignment_text_horizontal, vertical=alignment_text_vertical, wrapText = True)
			book.cell(1, column = 1).border = border_all
			if header_height:
				book.row_dimensions[1].height = header_height

		def column_name(column):
			book.cell(row_mini, column = column, value = str(name))
			book.cell(row_mini, column = column).font = styles.Font(color=styles.colors.WHITE, bold=True, size = font_size_name)
			book.cell(row_mini, column = column).fill = styles.PatternFill("solid", fgColor = "0099ff")
			book.cell(row_mini, column = column).alignment = styles.Alignment(horizontal=alignment_name_horizontal, vertical=alignment_name_vertical, wrapText = True)
			book.cell(row_mini, column = column).border = border_top_left_bottom

		def column_code(column):
			book.cell(row_mini, column = column, value = CODE + str(code))
			book.cell(row_mini, column = column).font = styles.Font(bold=True, size = font_size_code)
			book.cell(row_mini, column = column).fill = styles.PatternFill("solid", fgColor = "ffff00")
			book.cell(row_mini, column = column).alignment = styles.Alignment(horizontal=alignment_code_horizontal, vertical=alignment_code_vertical, wrapText = True, indent=1)
			book.cell(row_mini, column = column).border = border_top_bottom

		def column_price(column):
			book.cell(row_mini, column = column, value = PRICE + str(price))
			book.cell(row_mini, column = column).font = styles.Font(color=styles.colors.RED, bold=True, size = font_size_price)
			book.cell(row_mini, column = column).alignment = styles.Alignment(horizontal=alignment_code_horizontal, vertical=alignment_code_vertical, wrapText = True, indent=1)
			book.cell(row_mini, column = column).border = border_top_right_bottom

		def column_place(column):
			book.cell(row_mini, column = column).border = border_top_right_bottom

		# Перераховуємо функції в якому порядку яку інформацію розміщати в порядку по колонках
		column_order = [
			column_name,
			column_code,
			column_price,
		]

		# Перераховуємо для кожної колонки її ширину
		columns_width = [20, 20, 20]

		column_count = range(len(column_order))
		# Встановлюємо ширину колонки
		for num in column_count:
			index = num + 1
			column_letter = utils.cell.get_column_letter(index)
			book.column_dimensions[column_letter].width = columns_width[num]

		header()

		for doc in data:
			name = doc.get("name")
			code = doc.get("code")
			price = doc.get("price")

			for column_function in column_order:
				column_function(column)
				column += 1

			if row_height:
				book.row_dimensions[row_mini].height = row_height
			column = 1
			row_mini += 1

		# Зберігаємо ексель книгу
		helps.saveBook(path_to_this_file, wb, file_name)

helps = tools.Showcase_tools()

# Номер групи ""
id_of_group = ""

temp = helps.get_available_goods_from_group(id_of_group)
temp = helps.sort_array(temp, "price", True)
print(temp)

create_workbook_list(__file__, helps, )

cwd = os.path.abspath(__file__)
helps.zip_archive(cwd)
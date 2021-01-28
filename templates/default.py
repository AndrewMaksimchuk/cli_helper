# 
# 
import os
from openpyxl import *
from excel_styles import *
from create_excel_file_templates import *
import tools

def create_workbook_(data,
													file_name,
													column_number = 4,
													labels_width = 21):
	wb = Workbook()
	ws = wb.create_sheet(file_name, 0)
	book = wb[file_name]

	column_count = range(column_number)
	# Встановлюємо ширину колонки
	for num in column_count:
		index = num + 1
		column_letter = utils.cell.get_column_letter(index)
		book.column_dimensions[column_letter].width = labels_width

	row_mini = 1
	column = 1
	count_cell = 0

	for doc in data:
		name = doc.get("name")
		code = doc.get("code")
		price = doc.get("price")

		book.cell(row_mini, column = column, value = name)
		book.cell(row_mini, column = column).fill = background_color_name
		book.cell(row_mini, column = column).font = font_name
		book.cell(row_mini, column = column).border = border_top
		row_mini += 1

		book.cell(row_mini, column = column, value = "Код: " + str(code))
		book.cell(row_mini, column = column).fill = background_color_code
		book.cell(row_mini, column = column).border = border_side
		row_mini += 1
		
		book.cell(row_mini, column = column, value = "Ціна: " + str(price) + " грн")
		book.cell(row_mini, column = column).font = font_price
		book.cell(row_mini, column = column).border = border_bottom
		
		count_cell += 1
		column += 1

		if column == column_number + 1:
			column = 1
		
		if count_cell == column_number:
			row_mini += 1
			count_cell = 0
		else:
			row_mini -= 2

	helps.saveBook(__file__, wb, file_name)

helps = tools.Showcase_tools()
# Номер групи ""
id_of_group = ""

temp = helps.get_available_goods_from_group(id_of_group)
print(temp)

file_name = ""
create_workbook_(temp, file_name)

cwd = os.path.abspath(__file__)
helps.zip_archive(cwd)